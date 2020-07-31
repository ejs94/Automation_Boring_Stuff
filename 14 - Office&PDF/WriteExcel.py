import os
import openpyxl
os.chdir('.\\14 - Office&PDF')
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'].value == None
sheet['A1'].value = 42
sheet['A2'].value = 'Hello'
wb.save('example.xlsx')
sheet2 = wb.create_sheet()
sheet2.title = 'Minha planilha linda'
print(wb.sheetnames)
wb.save('example.xlsx')
wb.create_sheet(index=0, title='Minha Outra Planilha Linda')
wb.save('example.xlsx')
